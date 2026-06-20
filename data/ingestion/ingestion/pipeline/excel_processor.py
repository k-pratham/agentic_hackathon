"""
Excel KRI data processor — parses Excel files into kri_data records.

Uses openpyxl to read Excel files and maps columns based on the
configurable KriColumnMapping from settings.
"""

import logging
from datetime import datetime, date

import openpyxl

from ingestion.config.settings import KriColumnMapping

logger = logging.getLogger(__name__)


def parse_excel_to_kri(
    file_path: str, column_mapping: KriColumnMapping
) -> list[dict]:
    """
    Parse an Excel file into a list of KRI data records.

    Reads the first sheet, identifies columns by header name using the
    configurable column mapping, and extracts rows as dictionaries.

    Args:
        file_path: Path to the Excel (.xlsx) file.
        column_mapping: KriColumnMapping with expected column header names.

    Returns:
        List of dicts with keys: indicator_code, indicator_name,
        actual_value, threshold_limit, report_date.

    Raises:
        ValueError: If required columns are not found in the Excel file.
    """
    logger.info("Parsing Excel file for KRI data: %s", file_path)

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            raise ValueError(f"No active sheet found in {file_path}")

        # Read header row to build column index map
        headers: list[str] = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            headers.append(str(cell.value or "").strip())

        # Map expected column names to column indices
        required_columns = {
            "indicator_code": column_mapping.indicator_code,
            "indicator_name": column_mapping.indicator_name,
            "actual_value": column_mapping.actual_value,
            "report_date": column_mapping.report_date,
        }
        optional_columns = {
            "threshold_limit": column_mapping.threshold_limit,
        }

        col_indices: dict[str, int] = {}

        for field_name, header_name in required_columns.items():
            try:
                col_indices[field_name] = headers.index(header_name)
            except ValueError:
                raise ValueError(
                    f"Required column '{header_name}' (for {field_name}) "
                    f"not found in Excel headers: {headers}"
                )

        for field_name, header_name in optional_columns.items():
            try:
                col_indices[field_name] = headers.index(header_name)
            except ValueError:
                logger.warning(
                    "Optional column '%s' (for %s) not found, will use None",
                    header_name,
                    field_name,
                )
                col_indices[field_name] = -1

        # Parse data rows
        records: list[dict] = []
        row_count = 0

        for row in ws.iter_rows(min_row=2):
            row_count += 1
            cells = list(row)

            # Skip completely empty rows
            if all(c.value is None for c in cells):
                continue

            try:
                indicator_code = str(cells[col_indices["indicator_code"]].value or "").strip()
                indicator_name = str(cells[col_indices["indicator_name"]].value or "").strip()

                # Parse numeric value
                actual_raw = cells[col_indices["actual_value"]].value
                actual_value = _parse_numeric(actual_raw, "actual_value", row_count)

                # Parse optional threshold
                threshold_limit = None
                if col_indices["threshold_limit"] >= 0:
                    threshold_raw = cells[col_indices["threshold_limit"]].value
                    if threshold_raw is not None:
                        threshold_limit = _parse_numeric(
                            threshold_raw, "threshold_limit", row_count
                        )

                # Parse date
                report_date = _parse_date(
                    cells[col_indices["report_date"]].value, row_count
                )

                if not indicator_code or actual_value is None or report_date is None:
                    logger.warning("Skipping row %d: missing required fields", row_count + 1)
                    continue

                records.append({
                    "indicator_code": indicator_code,
                    "indicator_name": indicator_name,
                    "actual_value": actual_value,
                    "threshold_limit": threshold_limit,
                    "report_date": report_date,
                })

            except Exception:
                logger.error(
                    "Error parsing row %d in %s", row_count + 1, file_path,
                    exc_info=True,
                )
                continue

        logger.info(
            "Parsed %d KRI records from %d data rows in %s",
            len(records),
            row_count,
            file_path,
        )
        return records

    finally:
        wb.close()


def _parse_numeric(value, field_name: str, row_num: int) -> float | None:
    """Parse a numeric value from an Excel cell."""
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        logger.warning(
            "Non-numeric value '%s' in %s at row %d", value, field_name, row_num + 1
        )
        return None


def _parse_date(value, row_num: int) -> date | None:
    """Parse a date value from an Excel cell."""
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    # Try string parsing
    str_val = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(str_val, fmt).date()
        except ValueError:
            continue

    logger.warning("Cannot parse date '%s' at row %d", value, row_num + 1)
    return None


def parse_compliance_excel(file_path: str) -> list[dict]:
    """
    Parse a compliance Excel file into structured observations.

    Supports both standard horizontal (rows = records) and transposed
    (columns = records) layouts. Returns records with explicit structured
    fields mapped to the EsFields schema for precise agentic queries.

    Returns:
        List of dicts with keys:
          - supervised_entity_name (str)
          - sr_no (str)
          - action_para_id / observation_id (str)  — e.g. "3.1.1"
          - observation / observation_text (str)
          - actual_due_date (str)
          - extended_due_date (str)
          - compliance_status (str)  — "Not Complied", "Complied", etc.
          - se_status (str)          — entity's self-reported status
          - response_type (str)
          - is_resubmission_required (str)
          - raw_data (dict)
          - text_content (str)
    """
    logger.info("Parsing compliance Excel file: %s", file_path)
    wb = openpyxl.load_workbook(file_path, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            raise ValueError(f"No active sheet in compliance file: {file_path}")

        max_r = ws.max_row
        max_c = ws.max_column

        # Detect layout: transposed or horizontal
        col_a_vals = []
        for r in range(1, min(max_r + 1, 15)):
            val = ws.cell(row=r, column=1).value
            if val is not None:
                col_a_vals.append(str(val).strip().lower())

        transposed_indicators = [
            "supervised entity name", "action para id",
            "action para description", "se status", "sr.no."
        ]
        is_transposed = any(ind in col_a_vals for ind in transposed_indicators)

        raw_records: list[dict] = []
        if is_transposed:
            logger.info("Detected transposed layout for compliance sheet")
            keys = []
            for r in range(1, max_r + 1):
                val = ws.cell(row=r, column=1).value
                keys.append(str(val).strip() if val is not None else f"Field_{r}")
            for c in range(2, max_c + 1):
                col_vals = [ws.cell(row=r, column=c).value for r in range(1, max_r + 1)]
                if all(v is None for v in col_vals):
                    continue
                record = {}
                for r, key in enumerate(keys):
                    val = col_vals[r]
                    record[key] = str(val).strip() if val is not None else ""
                raw_records.append(record)
        else:
            logger.info("Detected standard horizontal layout for compliance sheet")
            headers = []
            for c in range(1, max_c + 1):
                val = ws.cell(row=1, column=c).value
                headers.append(str(val).strip() if val is not None else f"Header_{c}")
            for r in range(2, max_r + 1):
                row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_c + 1)]
                if all(v is None for v in row_vals):
                    continue
                record = {}
                for c, header in enumerate(headers):
                    val = row_vals[c]
                    record[header] = str(val).strip() if val is not None else ""
                raw_records.append(record)

        def _find_val(record: dict, keywords: list[str]) -> str:
            for kw in keywords:
                for k, v in record.items():
                    if kw.lower() in k.lower():
                        return v
            return ""

        parsed_records: list[dict] = []
        for rec in raw_records:
            try:
                supervised_entity = _find_val(rec, ["supervised entity", "banks name", "bank name"])
                sr_no = _find_val(rec, ["sr.no.", "sr number", "serial number"])
                action_para_id = _find_val(rec, ["action para id", "action para number", "para id"])
                observation = _find_val(rec, ["observation", "action para description"])
                actual_due_date = _find_val(rec, ["actual due date", "due date for compliance"])
                extended_due_date = _find_val(rec, ["extended due date", "extended due date requested"])
                response_type = _find_val(rec, ["response type"])
                is_resub = _find_val(rec, ["is resubmission", "resubmission required"])

                # Prefer explicit 'Status' column; fallback to any 'status' column
                status = _find_val(rec, ["status of observation as per the auditor"])
                if not status:
                    status = _find_val(rec, ["status"])

                se_status = _find_val(rec, ["se status", "se response status", "entity status"])

                text_parts = [
                    f"Supervised Entity: {supervised_entity}",
                    f"Action Para ID: {action_para_id}",
                    f"Observation: {observation}",
                    f"Due Date: {actual_due_date}",
                    f"Extended Due Date: {extended_due_date}",
                    f"Auditor Status (Compliance): {status}",
                    f"SE Status: {se_status}",
                ]

                comments_found = []
                for key, val in rec.items():
                    if val and any(word in key.lower() for word in
                                   ["comment", "response", "reply", "remark"]):
                        comments_found.append(f"{key}: {val}")
                if comments_found:
                    text_parts.append("\n--- Comments and Response Log ---\n" +
                                      "\n".join(comments_found))

                text_content = "\n".join(text_parts)

                parsed_records.append({
                    "supervised_entity_name": supervised_entity,
                    "sr_no": sr_no,
                    "action_para_id": action_para_id,
                    "observation_id": action_para_id,
                    "observation": observation,
                    "observation_text": observation,
                    "actual_due_date": actual_due_date,
                    "extended_due_date": extended_due_date,
                    "compliance_status": status,
                    "se_status": se_status,
                    "response_type": response_type,
                    "is_resubmission_required": is_resub,
                    "raw_data": rec,
                    "text_content": text_content,
                })
            except Exception as e:
                logger.error("Error parsing compliance record: %s", str(e), exc_info=True)
                continue

        logger.info("Extracted %d structured compliance records from Excel", len(parsed_records))
        return parsed_records

    finally:
        wb.close()


def parse_generic_excel_to_markdown(file_path: str) -> list[str]:
    """
    Parse generic Excel sheets and convert tables to Markdown strings.
    This preserves tabular layout for vector semantic search.
    """
    logger.info("Converting Excel sheets to markdown: %s", file_path)
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheets_md: list[str] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Skip empty sheets
            if ws.max_row <= 1 and ws.max_column <= 1 and ws.cell(row=1, column=1).value is None:
                continue

            md_lines = [f"### Sheet: {sheet_name}"]
            for r in range(1, ws.max_row + 1):
                row_vals = []
                for c in range(1, ws.max_column + 1):
                    val = ws.cell(row=r, column=c).value
                    row_vals.append(str(val).replace("\n", " ").strip() if val is not None else "")
                
                # Check if row is empty
                if all(v == "" for v in row_vals):
                    continue

                md_lines.append("| " + " | ".join(row_vals) + " |")
                if r == 1:
                    # Add divider
                    md_lines.append("| " + " | ".join(["---"] * len(row_vals)) + " |")

            sheet_md = "\n".join(md_lines).strip()
            if sheet_md:
                sheets_md.append(sheet_md)
        return sheets_md
    finally:
        wb.close()


def parse_kri_questionnaire_excel(file_path: str) -> list[dict]:
    """
    Parse a KRI questionnaire excel file.
    
    Reads rows containing question numbers, question text, and responses,
    and returns a structured dict for each question suitable for Elasticsearch indexing.
    """
    logger.info("Parsing KRI questionnaire Excel file: %s", file_path)
    wb = openpyxl.load_workbook(file_path, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            raise ValueError(f"No active sheet in KRI file: {file_path}")

        max_r = ws.max_row
        max_c = ws.max_column

        # Read row 1 as headers
        headers = []
        for c in range(1, max_c + 1):
            val = ws.cell(row=1, column=c).value
            headers.append(str(val).strip() if val is not None else f"Header_{c}")

        records: list[dict] = []
        for r in range(2, max_r + 1):
            try:
                row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_c + 1)]
                if all(v is None for v in row_vals):
                    continue

                rec = {}
                for c, header in enumerate(headers):
                    val = row_vals[c]
                    rec[header] = str(val).strip() if val is not None else ""

                # Check if it has a Question/Question Number to skip group headers/empty items
                def find_val(keywords: list[str]) -> str:
                    for kw in keywords:
                        for k, v in rec.items():
                            if kw.lower() in k.lower():
                                return v
                    return ""

                question_number = find_val(["question number", "q.no", "number"])
                question_text = find_val(["question", "indicator name", "description"])
                response = find_val(["response", "actual value", "answer"])
                se_name = find_val(["se name", "bank name", "supervised entity name"])
                se_type = find_val(["se type", "entity type"])
                response_code = find_val(["response code", "indicator code"])
                response_status = find_val(["response status", "status"])

                # Skip header rows inside data rows (like "T1 Question Group" with empty response)
                if not question_number or not question_text:
                    continue

                text_parts = [
                    f"Question Number: {question_number}",
                    f"Question: {question_text}",
                    f"Response: {response}",
                    f"Response Code: {response_code}",
                    f"SE Name: {se_name}",
                ]
                if se_type:
                    text_parts.append(f"SE Type: {se_type}")
                if response_status:
                    text_parts.append(f"Response Status: {response_status}")

                text_content = "\n".join(text_parts)

                records.append({
                    "question_number": question_number,
                    "question": question_text,
                    "response": response,
                    "se_name": se_name,
                    "response_code": response_code,
                    "text_content": text_content
                })
            except Exception as e:
                logger.error(
                    "Error parsing KRI questionnaire row %d: %s",
                    r,
                    str(e),
                    exc_info=True
                )
                continue

        logger.info("Extracted %d KRI questionnaire records", len(records))
        return records

    finally:
        wb.close()
